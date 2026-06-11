"""Smoke-test core logic from app.py without needing streamlit installed."""
import json
import re
import statistics  # noqa: F401  (used by exec'd code)
import pandas as pd
from datetime import datetime
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

SRC = open("app.py", encoding="utf-8").read()

# Pull out only the pure functions (skip the st.cache_data-decorated scraper).
WANTED = ["slugify", "to_url", "room_label", "find_listing_dicts",
          "listing_link", "parse_page", "fair_price", "summarize",
          "build_excel", "advise"]
ns = {"json": json, "re": re, "statistics": statistics, "pd": pd,
      "BASE": "https://speedhome.com",
      "FURNISH_MAP": {"FULL": "Fully Furnished", "PARTIAL": "Partially Furnished",
                      "NONE": "Unfurnished"}}
from io import BytesIO  # noqa: E402
ns["BytesIO"] = BytesIO
for k, v in [("datetime", datetime), ("BarChart", BarChart), ("Reference", Reference), ("Alignment", Alignment), ("Font", Font), ("PatternFill", PatternFill), ("Border", Border), ("Side", Side), ("get_column_letter", get_column_letter), ("ArrayFormula", ArrayFormula), ("BRAND", "C8102E"), ("ARIAL", "Arial")]:
    ns[k] = v

for name in WANTED:
    m = re.search(rf"\ndef {name}\(.*?(?=\n\ndef |\n\n@|\n\n# )", SRC, re.S)
    assert m, f"could not extract {name}"
    exec(m.group(0), ns)

# ---- build mock __NEXT_DATA__ matching the probe_report structure ----
listings = [
    {"name": "Arte Mont Kiara", "title": "Cozy studio at Arte Mont Kiara",
     "furnishType": "FULL", "roomType": None, "sqft": 500, "bedroom": 0,
     "bathroom": 1, "carpark": 1, "price": 1800, "minRentalDuration": 12},
    {"name": "Kiara Kasih Condominium", "title": "Kiara Kasih 3BR corner",
     "furnishType": "PARTIAL", "roomType": None, "sqft": 1010, "bedroom": 3,
     "bathroom": 2, "carpark": 2, "price": 1920, "minRentalDuration": 12},
    {"name": "Duta Park Residences", "title": "Duta Park 2BR",
     "furnishType": "FULL", "roomType": None, "sqft": 772, "bedroom": 2,
     "bathroom": 2, "carpark": 1, "price": 1950, "minRentalDuration": 12},
    {"name": "Duta Park Residences", "title": "Duta Park 2BR high floor",
     "furnishType": "NONE", "roomType": None, "sqft": 800, "bedroom": 2,
     "bathroom": 2, "carpark": 1, "price": 2100, "minRentalDuration": 12},
    {"name": "Changkat View", "title": "Room at Changkat View",
     "furnishType": "FULL", "roomType": "MEDIUM", "sqft": 150, "bedroom": 1,
     "bathroom": 1, "carpark": 0, "price": 750, "minRentalDuration": 6},
    # outlier to test fair-price trimming
    {"name": "Luxury Penthouse", "title": "Penthouse 2BR",
     "furnishType": "FULL", "roomType": None, "sqft": 3000, "bedroom": 2,
     "bathroom": 3, "carpark": 3, "price": 15000, "minRentalDuration": 12},
]
next_data = {"props": {"pageProps": {"searchResult": {"items": listings}}}}
html = (
    '<html><a href="/details/arte-mont-kiara-xoehfyha">x</a>'
    '<a href="/details/kiara-kasih-condominium-mont-kiara-eothvqia">x</a>'
    '<a href="/details/duta-park-residences-afxprnpr">x</a>'
    '<a href="/details/changkat-view-dutamas-agkgqjvp">x</a>'
    f'<script id="__NEXT_DATA__" type="application/json">{json.dumps(next_data)}'
    '</script></html>'
)

rows, links = ns["parse_page"](html)
print(f"parsed {len(rows)} listings, {len(links)} detail links")
assert len(rows) == 6, rows

df = pd.DataFrame(rows)
print(df[["Title", "Room Type", "Price / month (RM)", "Price / year (RM)",
          "Furnishing", "Link"]].to_string(index=False))

assert set(df["Room Type"]) == {"Studio", "3BR", "2BR", "Room"}
assert (df["Price / year (RM)"] == df["Price / month (RM)"] * 12).all()
assert df.loc[0, "Link"].endswith("arte-mont-kiara-xoehfyha")
assert (df["Title"] != df["Property / Area"]).all()  # titles always distinct
assert "Fully Furnished" in set(df["Furnishing"])
assert "Unfurnished" in set(df["Furnishing"])

summary = ns["summarize"](df)
print("\n--- summary ---")
print(summary.to_string(index=False))
two_br = summary[summary["Room Type"] == "2BR"].iloc[0]
assert two_br["Units Found"] == 3
# fair price should not be dragged to 15000 by the outlier
assert two_br["Fair Price (RM)"] < 5000, two_br

xlsx = ns["build_excel"](summary, df, "Mont Kiara", "live", "test-url")
assert xlsx[:2] == b"PK" and len(xlsx) > 4000
print(f"\nexcel bytes: {len(xlsx):,} ✅")

assert ns["to_url"]("Mont Kiara") == "https://speedhome.com/rent/mont-kiara"
assert ns["to_url"]("https://speedhome.com/rent/bangsar") == \
    "https://speedhome.com/rent/bangsar"
top, info = ns["advise"](df, summary, 2500, [], [], "best")
assert not top.empty and not info["empty"] and info["count"] == 5
assert info["gap"] is not None and info["price"] <= 2500
top2, info2 = ns["advise"](df, summary, 100, [], [], "cheap")
assert top2.empty and info2["empty"] and info2["cheapest"] == 750
print("advisor OK: count =", info["count"], "| top =", info["title"])

print("\nALL TESTS PASSED ✅")
