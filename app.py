"""
J360 Rental Intelligence — SPEEDHOME Property Price Intelligence App
=====================================================================
Vibe Coding Test — CEO Office, Jendela360

Data source: public SPEEDHOME.com search pages. SPEEDHOME is a Next.js
site that embeds its full result set as JSON inside
<script id="__NEXT_DATA__">, giving clean per-listing data (price, sqft,
bedrooms, furnishing, links) without fragile HTML parsing.

Politeness: fetches and respects robots.txt, delays between requests.
Fallback: every successful scrape is snapshotted to data/*.json so the
deployed app still shows real data if the host's IP is blocked.
UI: bilingual (English / Bahasa Indonesia) via the 🌐 selector.
"""

import glob
import json
import os
import re
import statistics
import time
import urllib.robotparser
from datetime import datetime
from io import BytesIO

import cloudscraper
import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

# ----------------------------------------------------------------- constants

BASE = "https://speedhome.com"
DATA_DIR = "data"
os.makedirs(DATA_DIR, exist_ok=True)

BRAND = "C8102E"          # J360-style red used in headers
ARIAL = "Arial"

SEED_SUGGESTIONS = [
    "Mont Kiara", "Mont Kiara Aman", "Mont Kiara Bayu", "Mont Kiara Pines",
    "Bangsar", "Bangsar South", "KLCC", "KL City Centre", "Bukit Bintang",
    "Cheras", "Kepong", "Setapak", "Wangsa Maju", "Sentul", "Dutamas",
    "Sri Hartamas", "Desa ParkCity", "Taman Tun Dr Ismail", "Ampang",
    "Old Klang Road", "Sri Petaling", "Bukit Jalil", "Kuchai Lama",
    "Petaling Jaya", "Subang Jaya", "USJ", "Sunway", "Bandar Sunway",
    "Shah Alam", "Setia Alam", "Kota Damansara", "Damansara Perdana",
    "Mutiara Damansara", "Bandar Utama", "Ara Damansara", "Kelana Jaya",
    "Puchong", "Bandar Puteri Puchong", "Seri Kembangan", "Serdang",
    "Cyberjaya", "Putrajaya", "Kajang", "Semenyih", "Bangi", "Rawang",
    "Selayang", "Batu Caves", "Gombak", "Ampang Jaya", "Pandan Indah",
    "Johor Bahru", "Skudai", "Iskandar Puteri", "Penang", "George Town",
    "Bayan Lepas", "Butterworth", "Ipoh", "Seremban", "Melaka",
    "Kota Kinabalu", "Kuching",
]

FURNISH_MAP = {
    "FULL": "Fully Furnished",
    "PARTIAL": "Partially Furnished",
    "PARTIALLY": "Partially Furnished",
    "NONE": "Unfurnished",
    "UNFURNISHED": "Unfurnished",
}

BAD_SUGGESTION_TOKENS = (
    "wifi", "termasuk", "bilik", "kamar", "room", "include", "furnish",
    "private", "near ", "mrt", "lrt", "sewa", "rent", "deposit", "utilities",
)

# ------------------------------------------------------------ translations

T = {
 "en": {
  "caption": "Live property price intelligence from SPEEDHOME.com — "
             "search an area, get instant market statistics.",
  "tab_search": "🔍 Search area / property",
  "tab_url": "🔗 Paste URL",
  "tab_cmp": "⚖️ Compare areas",
  "search_label": "Type an area or property name "
                  "(suggestions appear as you type)",
  "search_ph": 'Try typing "Mont" …',
  "btn_analyze": "🚀 Analyze rentals",
  "btn_analyze_url": "🚀 Analyze URL",
  "url_label": "SPEEDHOME URL",
  "url_err": "Please paste a speedhome.com URL.",
  "url_first": "Paste a SPEEDHOME URL first.",
  "pick_first": "Pick an area first, then press Analyze.",
  "cmp_label": "Pick 2–3 areas to compare side by side",
  "btn_cmp": "⚖️ Compare areas",
  "cmp_caption": "First comparison takes up to a minute (polite delays "
                 "between requests). Results are cached for an hour.",
  "cmp_two": "Pick at least two areas to compare.",
  "cmp_fail": "Could not collect data for at least two areas — "
              "try again or pick different areas.",
  "settings": "⚙️ Scrape settings",
  "pages": "Max pages to fetch",
  "pages_help": "40 listings per page. More pages = slower.",
  "delay": "Delay between requests (seconds)",
  "delay_help": "Be polite to SPEEDHOME's servers.",
  "spinner": "Collecting live data for {} …",
  "no_results": "No listings found for this search and no snapshot "
                "available. Tip: property names work best without the "
                "city part; or try the area name instead. ",
  "snap_warn": "Live fetch returned no data (the host's IP may be "
               "blocked), so this shows the saved snapshot ({}) — real "
               "data previously scraped from SPEEDHOME.",
  "units_found": "units found",
  "m_units": "Units", "m_median": "Median rent", "m_cheap": "Cheapest",
  "m_exp": "Most expensive",
  "rent_types": "Rent types covered",
  "rt_month": "Monthly: {} listings",
  "rt_year": "Yearly: {} (= monthly × 12, shown per listing)",
  "rt_day": "Daily: not offered on SPEEDHOME for this area — SPEEDHOME "
            "focuses on monthly/yearly tenancy.",
  "sum_head": "📊 Price Summary by Unit Type",
  "fair_cap": "Fair Price = median after removing statistical outliers "
              "(IQR method) — a representative mid-market price.",
  "list_head": "🏠 Unit Listings",
  "f_room": "Room type", "f_furn": "Furnishing",
  "f_budget": "Budget / max price (RM/month)",
  "f_budget_help": "0 = no limit. Type a number, then press Enter "
                   "to apply.",
  "dl_xlsx": "⬇️ Download Excel (.xlsx)", "dl_csv": "⬇️ Download CSV",
  "adv_head": "🤖 Rental Advisor — find your best match",
  "adv_budget": "Your monthly budget (RM)",
  "adv_prio": "What matters most?",
  "adv_opts": ["Best deal vs market (recommended)", "Cheapest rent",
               "Biggest space", "Lowest RM/sqft"],
  "adv_types": "Room types you'd accept",
  "adv_furn": "Furnishing you'd accept",
  "adv_none": "No listings match. The cheapest unit in this data is "
              "RM {:,.0f}/month — try raising the budget or relaxing "
              "the filters.",
  "below": "{:.0f}% below", "above": "{:.0f}% above", "at": "exactly at",
  "adv_msg": "With a budget of RM {b:,.0f}, you have **{n} matching "
             "units**. Top pick: **{title}** at RM {p:,.0f}/month — "
             "{gap} the {room} fair price (RM {fair:,.0f})",
  "adv_sqft": ", {s:.0f} sqft at RM {psf:.2f}/sqft.",
  "bp_head": "✨ **Best pick: {}** — ",
  "bp_fair": "priced {:.0f}% {} the {} fair price",
  "bp_below": "below", "bp_above": "above",
  "bp_space": "more space than the segment average ({:.0f} vs "
              "{:.0f} sqft)",
  "bp_furn": "fully furnished, so no upfront furniture cost",
  "bp_year": "total yearly cost RM {:,.0f}",
  "vsfair_cap": "Negative 'vs Fair' = priced below the market's fair "
                "price for that room type — likely a good deal.",
  "roi_head": "🧮 ROI Calculator — for property investors",
  "roi_type": "Unit type to buy", "roi_price": "Purchase price (RM)",
  "roi_cost": "Monthly costs (RM)",
  "roi_cost_help": "Maintenance fee, sinking fund, insurance, etc.",
  "roi_rent": "Expected rent", "roi_rent_sub": "{} fair price in {}",
  "roi_gross": "Gross yield", "roi_net": "Net yield",
  "roi_payback": "Payback period", "roi_yrs": "yrs",
  "roi_never": "Never (costs > rent)",
  "roi_strong": "📈 Strong: net yield above the ~3–5% typical range for "
                "Klang Valley condos.",
  "roi_ok": "📊 Reasonable: net yield within the typical 3–5% range for "
            "Klang Valley condos.",
  "roi_low": "📉 Low: net yield below the typical 3–5% range — the "
             "price may be high relative to local rents.",
  "roi_cap": "Gross yield = fair-price rent × 12 ÷ purchase price. Net "
             "yield deducts monthly costs. Based on live {} rental data "
             "above — estimates only, not financial advice.",
  "roi_nodata": "Not enough rental data for this unit type to estimate "
                "ROI.",
  "ins_head": "📈 Market Insights",
  "ch_tabs": ["💰 Typical rent by room type", "📦 Price ranges",
              "📊 Where prices cluster"],
  "rm_month": "RM / month", "n_listings": "Number of listings",
  "median_lbl": "Median (RM)", "avg_lbl": "Average (RM)",
  "median_line": "Median RM {:,.0f}",
  "ch1_cap": "Median = typical price (half cost more, half cost less). "
             "If Average is far above Median, a few expensive units are "
             "pulling it up.",
  "ch2_cap": "Each dot is one real listing. The box shows where the "
             "middle 50% of prices sit; dots far above are premium "
             "outliers.",
  "ch3_cap": "Taller bars = more listings at that price level. The "
             "dashed line marks the area's median rent.",
  "auto_ins": "**🤖 Auto insights for {a}:** the most available unit "
              "type is **{rt}** ({n} units, fair price RM {f:,.0f}). "
              "The most affordable segment is **{c}** (average "
              "RM {avg:,.0f}). Overall median rent is "
              "**RM {m:,.0f}/month**",
  "auto_psf": ", or about **RM {p:.2f} per sqft**.",
  "diag": "🔧 Diagnostics",
  "cmp_head": "⚖️ Area comparison",
  "cmp_chart": "Median rent by room type",
  "cmp_cheaper": "**{t}**: {lo} is RM {d:,.0f}/month cheaper than {hi}",
  "cmp_dl": "⬇️ Download comparison CSV",
  "footer": "J360 Rental Intelligence · data from public SPEEDHOME.com "
            "pages · respects robots.txt · built with AI assistance "
            "(vibe coding)",
 },
 "id": {
  "caption": "Intelijen harga properti langsung dari SPEEDHOME.com — "
             "cari area, dapatkan statistik pasar secara instan.",
  "tab_search": "🔍 Cari area / properti",
  "tab_url": "🔗 Tempel URL",
  "tab_cmp": "⚖️ Bandingkan area",
  "search_label": "Ketik nama area atau properti "
                  "(saran muncul saat Anda mengetik)",
  "search_ph": 'Coba ketik "Mont" …',
  "btn_analyze": "🚀 Analisis sewa",
  "btn_analyze_url": "🚀 Analisis URL",
  "url_label": "URL SPEEDHOME",
  "url_err": "Harap tempel URL speedhome.com.",
  "url_first": "Tempel URL SPEEDHOME terlebih dahulu.",
  "pick_first": "Pilih area dulu, lalu tekan Analisis.",
  "cmp_label": "Pilih 2–3 area untuk dibandingkan berdampingan",
  "btn_cmp": "⚖️ Bandingkan area",
  "cmp_caption": "Perbandingan pertama memakan waktu hingga satu menit "
                 "(jeda sopan antar permintaan). Hasil di-cache satu jam.",
  "cmp_two": "Pilih minimal dua area untuk dibandingkan.",
  "cmp_fail": "Tidak dapat mengumpulkan data untuk minimal dua area — "
              "coba lagi atau pilih area lain.",
  "settings": "⚙️ Pengaturan pengambilan data",
  "pages": "Jumlah halaman maksimum",
  "pages_help": "40 listing per halaman. Lebih banyak = lebih lambat.",
  "delay": "Jeda antar permintaan (detik)",
  "delay_help": "Bersikap sopan ke server SPEEDHOME.",
  "spinner": "Mengumpulkan data langsung untuk {} …",
  "no_results": "Tidak ada listing ditemukan dan tidak ada snapshot "
                "tersimpan. Tips: nama properti tanpa nama kota biasanya "
                "lebih akurat; atau coba nama area. ",
  "snap_warn": "Pengambilan langsung tidak mengembalikan data (IP server "
               "mungkin diblokir), jadi ditampilkan snapshot tersimpan "
               "({}) — data asli hasil scraping SPEEDHOME sebelumnya.",
  "units_found": "unit ditemukan",
  "m_units": "Unit", "m_median": "Sewa median", "m_cheap": "Termurah",
  "m_exp": "Termahal",
  "rent_types": "Tipe sewa yang dicakup",
  "rt_month": "Bulanan: {} listing",
  "rt_year": "Tahunan: {} (= bulanan × 12, ditampilkan per listing)",
  "rt_day": "Harian: tidak tersedia di SPEEDHOME untuk area ini — "
            "SPEEDHOME fokus pada sewa bulanan/tahunan.",
  "sum_head": "📊 Ringkasan Harga per Tipe Unit",
  "fair_cap": "Harga Wajar (Fair Price) = median setelah outlier "
              "statistik dihapus (metode IQR) — harga pasar menengah "
              "yang representatif.",
  "list_head": "🏠 Daftar Unit",
  "f_room": "Tipe kamar", "f_furn": "Perabotan",
  "f_budget": "Anggaran / harga maks (RM/bulan)",
  "f_budget_help": "0 = tanpa batas. Ketik angka, lalu tekan Enter.",
  "dl_xlsx": "⬇️ Unduh Excel (.xlsx)", "dl_csv": "⬇️ Unduh CSV",
  "adv_head": "🤖 Penasihat Sewa — temukan yang paling cocok untuk Anda",
  "adv_budget": "Anggaran bulanan Anda (RM)",
  "adv_prio": "Apa yang paling penting?",
  "adv_opts": ["Penawaran terbaik vs pasar (direkomendasikan)",
               "Sewa termurah", "Ruang terluas", "RM/sqft terendah"],
  "adv_types": "Tipe kamar yang Anda terima",
  "adv_furn": "Perabotan yang Anda terima",
  "adv_none": "Tidak ada yang cocok. Unit termurah di data ini "
              "RM {:,.0f}/bulan — naikkan anggaran atau longgarkan "
              "filter.",
  "below": "{:.0f}% di bawah", "above": "{:.0f}% di atas",
  "at": "tepat pada",
  "adv_msg": "Dengan anggaran RM {b:,.0f}, ada **{n} unit yang cocok**. "
             "Pilihan teratas: **{title}** seharga RM {p:,.0f}/bulan — "
             "{gap} harga wajar {room} (RM {fair:,.0f})",
  "adv_sqft": ", {s:.0f} sqft dengan RM {psf:.2f}/sqft.",
  "bp_head": "✨ **Pilihan terbaik: {}** — ",
  "bp_fair": "harga {:.0f}% {} harga wajar {}",
  "bp_below": "di bawah", "bp_above": "di atas",
  "bp_space": "lebih luas dari rata-rata segmen ({:.0f} vs {:.0f} sqft)",
  "bp_furn": "perabotan lengkap, tanpa biaya furnitur di awal",
  "bp_year": "total biaya tahunan RM {:,.0f}",
  "vsfair_cap": "Nilai 'vs Fair' negatif = harga di bawah harga wajar "
                "pasar untuk tipe kamar tersebut — kemungkinan "
                "penawaran bagus.",
  "roi_head": "🧮 Kalkulator ROI — untuk investor properti",
  "roi_type": "Tipe unit yang dibeli", "roi_price": "Harga beli (RM)",
  "roi_cost": "Biaya bulanan (RM)",
  "roi_cost_help": "Biaya maintenance, sinking fund, asuransi, dll.",
  "roi_rent": "Perkiraan sewa", "roi_rent_sub": "harga wajar {} di {}",
  "roi_gross": "Yield kotor", "roi_net": "Yield bersih",
  "roi_payback": "Periode balik modal", "roi_yrs": "thn",
  "roi_never": "Tidak pernah (biaya > sewa)",
  "roi_strong": "📈 Kuat: yield bersih di atas kisaran tipikal ~3–5% "
                "untuk kondominium Klang Valley.",
  "roi_ok": "📊 Wajar: yield bersih dalam kisaran tipikal 3–5% untuk "
            "kondominium Klang Valley.",
  "roi_low": "📉 Rendah: yield bersih di bawah kisaran tipikal 3–5% — "
             "harga mungkin terlalu tinggi dibanding sewa setempat.",
  "roi_cap": "Yield kotor = sewa harga wajar × 12 ÷ harga beli. Yield "
             "bersih dikurangi biaya bulanan. Berdasarkan data sewa {} "
             "di atas — hanya estimasi, bukan nasihat keuangan.",
  "roi_nodata": "Data sewa tipe unit ini belum cukup untuk estimasi ROI.",
  "ins_head": "📈 Wawasan Pasar",
  "ch_tabs": ["💰 Sewa tipikal per tipe kamar", "📦 Rentang harga",
              "📊 Sebaran harga"],
  "rm_month": "RM / bulan", "n_listings": "Jumlah listing",
  "median_lbl": "Median (RM)", "avg_lbl": "Rata-rata (RM)",
  "median_line": "Median RM {:,.0f}",
  "ch1_cap": "Median = harga tipikal (separuh lebih mahal, separuh "
             "lebih murah). Jika Rata-rata jauh di atas Median, ada "
             "beberapa unit mahal yang menariknya ke atas.",
  "ch2_cap": "Setiap titik adalah satu listing nyata. Kotak menunjukkan "
             "posisi 50% harga di tengah; titik jauh di atas adalah "
             "outlier premium.",
  "ch3_cap": "Batang lebih tinggi = lebih banyak listing di tingkat "
             "harga itu. Garis putus-putus menandai sewa median area.",
  "auto_ins": "**🤖 Wawasan otomatis untuk {a}:** tipe unit paling "
              "banyak tersedia adalah **{rt}** ({n} unit, harga wajar "
              "RM {f:,.0f}). Segmen paling terjangkau adalah **{c}** "
              "(rata-rata RM {avg:,.0f}). Median sewa keseluruhan "
              "**RM {m:,.0f}/bulan**",
  "auto_psf": ", atau sekitar **RM {p:.2f} per sqft**.",
  "diag": "🔧 Diagnostik",
  "cmp_head": "⚖️ Perbandingan area",
  "cmp_chart": "Sewa median per tipe kamar",
  "cmp_cheaper": "**{t}**: {lo} lebih murah RM {d:,.0f}/bulan daripada "
                 "{hi}",
  "cmp_dl": "⬇️ Unduh CSV perbandingan",
  "footer": "J360 Rental Intelligence · data dari halaman publik "
            "SPEEDHOME.com · menghormati robots.txt · dibangun dengan "
            "bantuan AI (vibe coding)",
 },
}

# ----------------------------------------------------------------- utilities


def slugify(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")


def to_url(query: str) -> str:
    """Turn user input (URL or area/property name) into a search URL.
    For names like 'Vipod Residence, Kuala Lumpur' only the part before
    the comma is used — that is the property slug SPEEDHOME expects."""
    q = query.strip()
    if q.startswith("http"):
        return q.split("#")[0]
    return f"{BASE}/rent/{slugify(q.split(',')[0])}"


def robots_allows(url: str, scraper) -> tuple[bool, str]:
    """Fetch robots.txt via cloudscraper (plain urllib gets blocked by
    Cloudflare with a 403, which Python's parser treats as disallow-all)."""
    try:
        resp = scraper.get(f"{BASE}/robots.txt", timeout=15)
        if resp.status_code != 200 or "<html" in resp.text[:200].lower():
            return True, (f"robots.txt not readable (HTTP {resp.status_code})"
                          " — proceeding carefully with delays")
        rp = urllib.robotparser.RobotFileParser()
        rp.parse(resp.text.splitlines())
        ok = rp.can_fetch("*", url)
        return ok, ("robots.txt allows this path ✅" if ok
                    else "robots.txt disallows this path ❌ — scraping aborted")
    except Exception as exc:
        return True, f"Could not read robots.txt ({exc}) — proceeding carefully"


def room_label(item: dict) -> str:
    if item.get("roomType"):
        return "Room"
    bed = item.get("bedroom")
    if bed in (0, None):
        return "Studio"
    try:
        return f"{int(bed)}BR"
    except (TypeError, ValueError):
        return "Unknown"


def find_listing_dicts(obj, found):
    """Recursively collect dicts that look like SPEEDHOME listings."""
    if isinstance(obj, dict):
        if "price" in obj and "sqft" in obj and "bedroom" in obj:
            found.append(obj)
        else:
            for v in obj.values():
                find_listing_dicts(v, found)
    elif isinstance(obj, list):
        for v in obj:
            find_listing_dicts(v, found)


def listing_link(item: dict, detail_links: list[str]) -> str:
    """Best-effort link: explicit slug/url field, else fuzzy match on name."""
    for key in ("seoUrl", "url", "slug", "detailUrl", "path"):
        val = item.get(key)
        if isinstance(val, str) and val:
            return val if val.startswith("http") else BASE + (
                val if val.startswith("/") else f"/details/{val}")
    name_slug = slugify(str(item.get("name") or item.get("title") or ""))
    if name_slug:
        tokens = [t for t in name_slug.split("-") if len(t) > 2][:3]
        for link in detail_links:
            if all(tok in link for tok in tokens):
                return BASE + link
    return ""


# ----------------------------------------------------------------- scraping


def parse_page(html: str) -> tuple[list[dict], list[str]]:
    """Extract listings from the __NEXT_DATA__ JSON of one page."""
    m = re.search(
        r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.S)
    if not m:
        return [], []
    try:
        data = json.loads(m.group(1))
    except json.JSONDecodeError:
        return [], []
    raw: list[dict] = []
    find_listing_dicts(data, raw)
    detail_links = sorted(set(re.findall(r'href="(/details/[^"]+)"', html))) \
        or sorted(set(re.findall(r'"(/details/[^"]+)"', html)))

    rows = []
    for it in raw:
        price = it.get("price")
        if not isinstance(price, (int, float)) or price <= 0:
            continue
        title = str(it.get("title") or it.get("name") or
                    it.get("propertyName") or "Listing").strip()
        prop = str(it.get("name") or it.get("propertyName") or
                   it.get("buildingName") or title).strip()
        rt = room_label(it)
        if title == prop:                    # search pages lack a separate
            title = f"{rt} · {prop}"         # title, so synthesize one
        furn = FURNISH_MAP.get(
            str(it.get("furnishType") or "").upper(), "Unknown")
        rows.append({
            "Title": title,
            "Property / Area": prop,
            "Room Type": rt,
            "Bedrooms": it.get("bedroom"),
            "Bathrooms": it.get("bathroom"),
            "Price / month (RM)": float(price),
            "Price / year (RM)": float(price) * 12,
            "Size (sqft)": it.get("sqft") or None,
            "Furnishing": furn,
            "Rent Type": "Monthly",      # SPEEDHOME lists monthly rentals
            "Min Duration (months)": it.get("minRentalDuration"),
            "Link": listing_link(it, detail_links),
        })
    return rows, detail_links


@st.cache_data(ttl=3600, show_spinner=False)
def scrape_speedhome(url: str, max_pages: int, delay: float) -> dict:
    """Scrape one search URL (cached 1h). Returns listings + diagnostics."""
    scraper = cloudscraper.create_scraper()
    ok, robots_msg = robots_allows(url, scraper)
    out = {"rows": [], "robots": robots_msg, "pages": 0, "errors": []}
    if not ok:
        return out
    seen = set()
    sep = "&" if "?" in url else "?"
    for page in range(1, max_pages + 1):
        page_url = url if page == 1 else f"{url}{sep}page={page}"
        try:
            resp = scraper.get(page_url, timeout=25)
            if resp.status_code != 200:
                out["errors"].append(f"Page {page}: HTTP {resp.status_code}")
                break
            rows, _ = parse_page(resp.text)
            fresh = [r for r in rows
                     if (r["Title"], r["Price / month (RM)"], r["Link"])
                     not in seen]
            for r in fresh:
                seen.add((r["Title"], r["Price / month (RM)"], r["Link"]))
            out["rows"].extend(fresh)
            out["pages"] = page
            if not fresh:                       # no new data → stop
                break
        except Exception as exc:
            out["errors"].append(f"Page {page}: {exc}")
            break
        if page < max_pages:
            time.sleep(delay)                   # polite delay
    return out


# ------------------------------------------------------------- snapshots


def snapshot_path(area: str) -> str:
    return os.path.join(DATA_DIR, f"snapshot_{slugify(area)}.json")


def save_snapshot(area: str, rows: list[dict]):
    with open(snapshot_path(area), "w", encoding="utf-8") as f:
        json.dump({"area": area, "saved": datetime.now().isoformat(),
                   "rows": rows}, f)


def load_snapshot(area: str) -> dict | None:
    path = snapshot_path(area)
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    candidates = glob.glob(os.path.join(DATA_DIR, "snapshot_*.json"))
    if candidates:
        with open(max(candidates, key=os.path.getmtime),
                  encoding="utf-8") as f:
            return json.load(f)
    return None


def valid_suggestion(name: str) -> bool:
    """Keep clean area/property names; reject room-advert sentences that
    SPEEDHOME stores as 'name' on room listings."""
    n = (name or "").strip()
    return (3 < len(n) <= 42
            and not any(ch in n for ch in "[]()!&/")
            and len(n.split()) <= 5
            and not any(tok in n.lower() for tok in BAD_SUGGESTION_TOKENS))


def load_suggestions() -> list[str]:
    path = os.path.join(DATA_DIR, "suggestions.json")
    extra = []
    if os.path.exists(path):
        try:
            extra = json.load(open(path, encoding="utf-8"))
        except Exception:
            extra = []
    pool = set(SEED_SUGGESTIONS) | set(extra)
    return sorted(n for n in pool if valid_suggestion(n))


def grow_suggestions(names: list[str]):
    path = os.path.join(DATA_DIR, "suggestions.json")
    current = set(load_suggestions())
    current |= {n.strip() for n in names if valid_suggestion(n)}
    json.dump(sorted(current), open(path, "w", encoding="utf-8"))


# --------------------------------------------------------------- statistics


def fair_price(values: list[float]) -> float | None:
    """Median of values inside the IQR fence (outliers removed)."""
    if not values:
        return None
    if len(values) < 4:
        return round(statistics.median(values), 2)
    s = sorted(values)
    q1 = s[len(s) // 4]
    q3 = s[(3 * len(s)) // 4]
    iqr = q3 - q1
    inside = [v for v in s if q1 - 1.5 * iqr <= v <= q3 + 1.5 * iqr]
    return round(statistics.median(inside or s), 2)


def summarize(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    order = {"Room": -1, "Studio": 0}
    for rt, grp in df.groupby("Room Type"):
        prices = grp["Price / month (RM)"].dropna().tolist()
        sizes = grp["Size (sqft)"].dropna().tolist()
        if not prices:
            continue
        try:
            mode_val = statistics.mode(prices)
        except statistics.StatisticsError:
            mode_val = statistics.median(prices)
        rows.append({
            "Room Type": rt,
            "Units Found": len(grp),
            "Average (RM)": round(statistics.mean(prices), 0),
            "Median (RM)": round(statistics.median(prices), 0),
            "Mode (RM)": round(mode_val, 0),
            "Fair Price (RM)": fair_price(prices),
            "Avg Size (sqft)": round(statistics.mean(sizes), 0)
            if sizes else None,
        })
    rows.sort(key=lambda r: order.get(
        r["Room Type"], int(re.sub(r"\D", "", r["Room Type"]) or 99)))
    return pd.DataFrame(rows)


# --------------------------------------------------------- excel export


def build_excel(summary: pd.DataFrame, listings: pd.DataFrame,
                area: str = "", source: str = "live",
                url: str = "") -> bytes:
    """Professional workbook: styled + bordered + banded sheets, live Excel
    formulas (yearly price, per-segment stats), hyperlinks, autofilter and
    a native chart."""
    buf = BytesIO()
    head_fill = PatternFill("solid", start_color=BRAND)
    band_fill = PatternFill("solid", start_color="F7F1F1")
    head_font = Font(name=ARIAL, bold=True, color="FFFFFF")
    body_font = Font(name=ARIAL, size=10)
    thin = Side(style="thin", color="D9D9D9")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    rm_fmt = '"RM" #,##0;[Red]("RM" #,##0);"-"'

    with pd.ExcelWriter(buf, engine="openpyxl") as xl:
        # ---------------- Listings sheet (data + formulas) ----------------
        cols = ["Title", "Property / Area", "Room Type", "Bedrooms",
                "Bathrooms", "Price / month (RM)", "Price / year (RM)",
                "Size (sqft)", "Furnishing", "Rent Type",
                "Min Duration (months)", "Link"]
        ldf = listings[cols].copy()
        ldf.to_excel(xl, sheet_name="Unit Listings", index=False)
        ws = xl.book["Unit Listings"]
        last = ws.max_row
        for cell in ws[1]:
            cell.fill, cell.font = head_fill, head_font
            cell.alignment, cell.border = center, box
        widths = [36, 30, 10, 10, 11, 16, 16, 11, 19, 11, 13, 9]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for r in range(2, last + 1):
            ws.cell(r, 7).value = f"=F{r}*12"        # yearly = monthly × 12
            link_cell = ws.cell(r, 12)
            link = link_cell.value
            if link:
                link_cell.value = "Open"
                link_cell.hyperlink = link            # native hyperlink
            else:
                link_cell.value = "-"
            for c in range(1, 13):
                cell = ws.cell(r, c)
                cell.font = body_font
                cell.border = box
                cell.alignment = left if c in (1, 2, 9) else center
                if r % 2 == 0:
                    cell.fill = band_fill
            for c in (6, 7):
                ws.cell(r, c).number_format = rm_fmt
            if ws.cell(r, 12).hyperlink:
                ws.cell(r, 12).font = Font(name=ARIAL, size=10,
                                           color="0563C1",
                                           underline="single")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:L{last}"

        # ---------------- Summary sheet (formula-driven) ----------------
        ss = xl.book.create_sheet("Price Summary", 0)
        ss["A1"] = f"SPEEDHOME Price Intelligence — {area}".strip(" —")
        ss["A1"].font = Font(name=ARIAL, bold=True, size=14, color=BRAND)
        ss.merge_cells("A1:G1")
        ss["A2"] = (f"Generated {datetime.now():%d %b %Y %H:%M} · "
                    f"{len(ldf)} listings · source: {source} · {url}")
        ss["A2"].font = Font(name=ARIAL, size=9, color="666666")
        ss.merge_cells("A2:G2")

        headers = ["Room Type", "Units Found", "Average (RM)", "Median (RM)",
                   "Mode (RM)", "Fair Price (RM)*", "Avg Size (sqft)"]
        for c, h in enumerate(headers, 1):
            cell = ss.cell(4, c, h)
            cell.fill, cell.font = head_fill, head_font
            cell.alignment, cell.border = center, box
        L = f"'Unit Listings'!$C$2:$C${last}"
        P = f"'Unit Listings'!$F$2:$F${last}"
        S = f"'Unit Listings'!$H$2:$H${last}"
        r = 5
        for _, row in summary.iterrows():
            ss.cell(r, 1, row["Room Type"])
            ss.cell(r, 2, f'=COUNTIF({L},A{r})')
            ss.cell(r, 3, f'=IFERROR(ROUND(AVERAGEIF({L},A{r},{P}),0),"-")')
            ss.cell(r, 4).value = ArrayFormula(
                f"D{r}", f'=IFERROR(ROUND(MEDIAN(IF({L}=A{r},{P})),0),"-")')
            ss.cell(r, 5).value = ArrayFormula(
                f"E{r}",
                f'=IFERROR(ROUND(MODE.SNGL(IF({L}=A{r},{P})),0),D{r})')
            ss.cell(r, 6, row["Fair Price (RM)"])    # computed in-app (IQR)
            ss.cell(r, 7,
                    f'=IFERROR(ROUND(AVERAGEIFS({S},{L},A{r},{S},">0"),0),"-")')
            for c in range(1, 8):
                cell = ss.cell(r, c)
                cell.font = body_font
                cell.border = box
                cell.alignment = left if c == 1 else center
                if r % 2 == 0:
                    cell.fill = band_fill
            for c in (3, 4, 5, 6):
                ss.cell(r, c).number_format = rm_fmt
            r += 1
        note = ss.cell(r + 1, 1, "*Fair Price = median after removing "
                                 "statistical outliers (IQR method), "
                                 "computed by the app.")
        note.font = Font(name=ARIAL, size=9, italic=True, color="666666")
        ss.cell(r + 3, 1, "Total units").font = Font(name=ARIAL, bold=True)
        ss.cell(r + 3, 2, f"=COUNTA('Unit Listings'!$A$2:$A${last})")
        ss.cell(r + 4, 1, "Overall median rent").font = Font(name=ARIAL,
                                                             bold=True)
        ss.cell(r + 4, 2, f"=ROUND(MEDIAN({P}),0)").number_format = rm_fmt
        for i, w in enumerate([16, 13, 14, 14, 14, 16, 15], 1):
            ss.column_dimensions[get_column_letter(i)].width = w

        chart = BarChart()
        chart.title = "Average monthly rent by room type"
        chart.y_axis.title = "RM / month"
        chart.height, chart.width = 8, 16
        n = len(summary)
        chart.add_data(Reference(ss, min_col=3, min_row=4, max_row=4 + n),
                       titles_from_data=True)
        chart.set_categories(Reference(ss, min_col=1, min_row=5,
                                       max_row=4 + n))
        ss.add_chart(chart, f"A{r + 7}")
    return buf.getvalue()


# ----------------------------------------------------------- advisor


def advise(df: pd.DataFrame, summary: pd.DataFrame, budget: float,
           types: list[str], furnish: list[str], code: str
           ) -> tuple[pd.DataFrame, dict]:
    """Rank listings for the renter. `code` is language-independent:
    best | cheap | big | psf. Returns (top picks, info for messaging)."""
    fair = dict(zip(summary["Room Type"], summary["Fair Price (RM)"]))
    cand = df[df["Price / month (RM)"] <= budget].copy()
    if types:
        cand = cand[cand["Room Type"].isin(types)]
    if furnish:
        cand = cand[cand["Furnishing"].isin(furnish)]
    if cand.empty:
        return cand, {"empty": True,
                      "cheapest": float(df["Price / month (RM)"].min())}
    cand["Fair (RM)"] = cand["Room Type"].map(fair)
    cand["vs Fair (%)"] = ((cand["Price / month (RM)"] - cand["Fair (RM)"])
                           / cand["Fair (RM)"] * 100).round(1)
    cand["RM/sqft"] = (cand["Price / month (RM)"]
                       / cand["Size (sqft)"]).round(2)
    if code == "cheap":
        cand = cand.sort_values("Price / month (RM)")
    elif code == "big":
        cand = cand.sort_values("Size (sqft)", ascending=False)
    elif code == "psf":
        cand = cand.sort_values("RM/sqft", na_position="last")
    else:                                       # best deal vs market
        cand = cand.sort_values(["vs Fair (%)", "RM/sqft"],
                                na_position="last")
    top = cand.head(5)
    b = top.iloc[0]
    return top, {
        "empty": False, "count": len(cand), "title": b["Title"],
        "price": float(b["Price / month (RM)"]), "room": b["Room Type"],
        "fair": float(b["Fair (RM)"]) if pd.notna(b["Fair (RM)"]) else None,
        "gap": float(b["vs Fair (%)"]) if pd.notna(b["vs Fair (%)"])
        else None,
        "sqft": float(b["Size (sqft)"]) if pd.notna(b["Size (sqft)"])
        else None,
        "psf": float(b["RM/sqft"]) if pd.notna(b["RM/sqft"]) else None,
    }


# ------------------------------------------------------------------- UI

st.set_page_config(page_title="J360 Rental Intelligence",
                   page_icon="🏢", layout="wide")
st.markdown("""
<style>
  .block-container {padding-top: 1.2rem;}
  [data-testid="stMetricValue"] {font-size: 1.5rem;}
  @media (max-width: 640px){
    .block-container {padding-left: .6rem; padding-right: .6rem;}
  }
</style>""", unsafe_allow_html=True)

st.title("🏢 J360 Rental Intelligence")
lang = st.radio("🌐 Language / Bahasa", ["English", "Bahasa Indonesia"],
                horizontal=True, key="lang")
t = T["id" if lang.startswith("Bahasa") else "en"]
st.caption(t["caption"])

tab_search, tab_url, tab_cmp = st.tabs(
    [t["tab_search"], t["tab_url"], t["tab_cmp"]])

# Freeze the suggestion list for this session: if the options changed
# mid-session (a scrape discovers new property names), Streamlit would
# reset the selectbox and swallow the user's next selection.
if "suggestions" not in st.session_state:
    st.session_state["suggestions"] = load_suggestions()
SUGG = st.session_state["suggestions"]

# Forms: nothing reruns while you type or pick — only when you press a button.
with tab_search:
    with st.form("f_search", border=False):
        sel = st.selectbox(t["search_label"], options=SUGG, index=None,
                           placeholder=t["search_ph"])
        go_search = st.form_submit_button(
            t["btn_analyze"], type="primary", use_container_width=True)
with tab_url:
    with st.form("f_url", border=False):
        url_in = st.text_input(
            t["url_label"],
            placeholder="https://speedhome.com/rent/mont-kiara")
        go_url = st.form_submit_button(
            t["btn_analyze_url"], type="primary", use_container_width=True)
with tab_cmp:
    with st.form("f_cmp", border=False):
        cmp_areas = st.multiselect(t["cmp_label"], options=SUGG,
                                   max_selections=3)
        cmp_go = st.form_submit_button(
            t["btn_cmp"], type="primary", use_container_width=True)
    st.caption(t["cmp_caption"])

with st.expander(t["settings"]):
    max_pages = st.slider(t["pages"], 1, 10, 3, help=t["pages_help"])
    delay = st.slider(t["delay"], 1.0, 5.0, 2.0, 0.5, help=t["delay_help"])


def get_rows(area_or_url: str) -> tuple[list[dict], str, dict]:
    """Scrape with snapshot fallback. Returns (rows, source, meta)."""
    url = to_url(area_or_url)
    area = (area_or_url.split(",")[0].strip()
            if not area_or_url.startswith("http")
            else url.rstrip("/").split("/")[-1].replace("-", " ").title())
    result = scrape_speedhome(url, max_pages, delay)
    rows, source = result["rows"], "live"
    if rows:
        save_snapshot(area, rows)
        grow_suggestions([r["Property / Area"] for r in rows])
    else:
        snap = load_snapshot(area)
        if snap:
            rows = snap["rows"]
            source = f"snapshot {snap['saved'][:16]}"
    return rows, source, {"area": area, "url": url,
                          "robots": result["robots"],
                          "pages": result["pages"],
                          "errors": result["errors"]}


# ---- single-area analysis trigger ----
query = None
if go_search and sel:
    query = sel
elif go_search:
    st.warning(t["pick_first"])
if go_url and url_in.strip():
    if "speedhome.com" not in url_in:
        st.error(t["url_err"])
    else:
        query = url_in.strip()
elif go_url:
    st.warning(t["url_first"])

if query:
    with st.spinner(t["spinner"].format(query)):
        rows, source, meta = get_rows(query)
    if rows:
        st.session_state["res"] = {"rows": rows, "source": source, **meta}
        st.session_state.pop("cmp", None)
    else:
        st.session_state.pop("res", None)
        st.error(t["no_results"] + " ".join(meta["errors"]))

# ---- comparison trigger ----
if cmp_go and len(cmp_areas) >= 2:
    bundles, prog = {}, st.progress(0.0, "…")
    for i, a in enumerate(cmp_areas):
        prog.progress(i / len(cmp_areas), t["spinner"].format(a))
        rows, source, meta = get_rows(a)
        if rows:
            bundles[a] = {"rows": rows, "source": source}
        if i < len(cmp_areas) - 1:
            time.sleep(delay)                  # polite gap between areas
    prog.empty()
    if len(bundles) >= 2:
        st.session_state["cmp"] = bundles
        st.session_state.pop("res", None)
    else:
        st.error(t["cmp_fail"])
elif cmp_go:
    st.warning(t["cmp_two"])


# =================================================== single-area results
def show_results():
    """Results persist in session_state, so filter/advisor interactions
    rerun the script with fresh values without losing the data."""
    res = st.session_state.get("res")
    if not res:
        return
    area_name, source, url = res["area"], res["source"], res["url"]
    df = pd.DataFrame(res["rows"])
    st.caption(f"🤖 {res['robots']} · pages: {res['pages']}")
    if source != "live":
        st.warning(t["snap_warn"].format(source))

    st.subheader(f"📍 {area_name} — {len(df)} {t['units_found']} "
                 f"({source})")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric(t["m_units"], len(df))
    c2.metric(t["m_median"],
              f"RM {df['Price / month (RM)'].median():,.0f}")
    c3.metric(t["m_cheap"], f"RM {df['Price / month (RM)'].min():,.0f}")
    c4.metric(t["m_exp"], f"RM {df['Price / month (RM)'].max():,.0f}")

    st.markdown(f"#### {t['rent_types']}")
    cov1, cov2, cov3 = st.columns(3)
    monthly_n = (df["Rent Type"] == "Monthly").sum()
    cov1.success(t["rt_month"].format(monthly_n))
    cov2.success(t["rt_year"].format(monthly_n))
    cov3.info(t["rt_day"])

    st.markdown(f"### {t['sum_head']}")
    summary = summarize(df)
    st.dataframe(summary, use_container_width=True, hide_index=True)
    st.caption(t["fair_cap"])

    st.markdown(f"### {t['list_head']}")
    fc1, fc2, fc3 = st.columns(3)
    f_type = fc1.multiselect(t["f_room"],
                             sorted(df["Room Type"].unique()))
    f_furn = fc2.multiselect(t["f_furn"],
                             sorted(df["Furnishing"].unique()))
    f_max = fc3.number_input(t["f_budget"], min_value=0, value=0,
                             help=t["f_budget_help"])
    view = df.copy()
    if f_type:
        view = view[view["Room Type"].isin(f_type)]
    if f_furn:
        view = view[view["Furnishing"].isin(f_furn)]
    if f_max:
        view = view[view["Price / month (RM)"] <= f_max]

    st.dataframe(
        view, use_container_width=True, hide_index=True, height=420,
        column_config={
            "Link": st.column_config.LinkColumn(
                "Listing", display_text="Open ↗"),
            "Price / month (RM)": st.column_config.NumberColumn(
                format="RM %.0f"),
            "Price / year (RM)": st.column_config.NumberColumn(
                format="RM %.0f"),
        })

    stamp = datetime.now().strftime("%Y%m%d")
    fname = f"SPEEDHOME_{area_name.replace(' ', '_')}_{stamp}"
    d1, d2 = st.columns(2)
    d1.download_button(
        t["dl_xlsx"], build_excel(summary, df, area_name, source, url),
        f"{fname}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True)
    d2.download_button(t["dl_csv"], df.to_csv(index=False),
                       f"{fname}.csv", "text/csv", use_container_width=True)

    # ---- Rental Advisor ----
    st.markdown(f"### {t['adv_head']}")
    a1, a2 = st.columns(2)
    p_lo = max(int(df["Price / month (RM)"].min() // 50 * 50), 100)
    p_hi = int(-(-df["Price / month (RM)"].max() // 50) * 50)
    if p_hi <= p_lo:
        p_hi = p_lo + 50
    budget = a1.slider(
        t["adv_budget"], min_value=p_lo, max_value=p_hi,
        value=min(max(int(df["Price / month (RM)"].median()), p_lo), p_hi),
        step=50)
    prio_label = a2.selectbox(t["adv_prio"], t["adv_opts"])
    code = ["best", "cheap", "big", "psf"][t["adv_opts"].index(prio_label)]
    a3, a4 = st.columns(2)
    adv_types = a3.multiselect(t["adv_types"],
                               sorted(df["Room Type"].unique()))
    adv_furn = a4.multiselect(t["adv_furn"],
                              sorted(df["Furnishing"].unique()))
    top, info = advise(df, summary, budget, adv_types, adv_furn, code)
    if info["empty"]:
        st.info("💡 " + t["adv_none"].format(info["cheapest"]))
    else:
        gap = info["gap"]
        gap_txt = (t["at"] if gap in (None, 0)
                   else t["below"].format(abs(gap)) if gap < 0
                   else t["above"].format(gap))
        msg = t["adv_msg"].format(b=budget, n=info["count"],
                                  title=info["title"], p=info["price"],
                                  gap=gap_txt, room=info["room"],
                                  fair=info["fair"] or 0)
        if info["psf"]:
            msg += t["adv_sqft"].format(s=info["sqft"], psf=info["psf"])
        st.info("💡 " + msg)

        b = top.iloc[0]
        seg = summary.set_index("Room Type")
        seg_size = (seg.loc[b["Room Type"], "Avg Size (sqft)"]
                    if b["Room Type"] in seg.index else None)
        points = []
        if gap not in (None, 0):
            d = t["bp_below"] if gap < 0 else t["bp_above"]
            points.append(t["bp_fair"].format(abs(gap), d, b["Room Type"]))
        if (seg_size and pd.notna(b["Size (sqft)"])
                and b["Size (sqft)"] > seg_size):
            points.append(t["bp_space"].format(b["Size (sqft)"], seg_size))
        if b["Furnishing"] == "Fully Furnished":
            points.append(t["bp_furn"])
        points.append(t["bp_year"].format(b["Price / year (RM)"]))
        st.success(t["bp_head"].format(b["Title"]) + "; ".join(points) + ".")

        st.dataframe(
            top[["Title", "Room Type", "Price / month (RM)", "Size (sqft)",
                 "RM/sqft", "Fair (RM)", "vs Fair (%)", "Furnishing",
                 "Link"]],
            use_container_width=True, hide_index=True,
            column_config={
                "Link": st.column_config.LinkColumn(
                    "Listing", display_text="Open ↗"),
                "Price / month (RM)": st.column_config.NumberColumn(
                    format="RM %.0f"),
                "Fair (RM)": st.column_config.NumberColumn(
                    format="RM %.0f"),
                "vs Fair (%)": st.column_config.NumberColumn(
                    format="%.1f%%"),
            })
        st.caption(t["vsfair_cap"])

    # ---- ROI calculator (investor view) ----
    st.markdown(f"### {t['roi_head']}")
    r1, r2, r3 = st.columns(3)
    roi_type = r1.selectbox(t["roi_type"], summary["Room Type"].tolist())
    buy_price = r2.number_input(t["roi_price"], min_value=50_000,
                                max_value=10_000_000, value=450_000,
                                step=10_000)
    monthly_cost = r3.number_input(t["roi_cost"], min_value=0,
                                   max_value=10_000, value=300, step=50,
                                   help=t["roi_cost_help"])
    fair_m = (summary.set_index("Room Type")
              .loc[roi_type, "Fair Price (RM)"])
    if pd.notna(fair_m) and fair_m > 0:
        gross = fair_m * 12 / buy_price * 100
        net_rent = fair_m - monthly_cost
        net = net_rent * 12 / buy_price * 100
        m1, m2, m3, m4 = st.columns(4)
        m1.metric(t["roi_rent"], f"RM {fair_m:,.0f}/mo",
                  t["roi_rent_sub"].format(roi_type, area_name),
                  delta_color="off")
        m2.metric(t["roi_gross"], f"{gross:.2f}% /yr")
        m3.metric(t["roi_net"], f"{net:.2f}% /yr")
        m4.metric(t["roi_payback"],
                  f"{buy_price / (net_rent * 12):.1f} {t['roi_yrs']}"
                  if net_rent > 0 else t["roi_never"])
        if net >= 5:
            st.success(t["roi_strong"])
        elif net >= 3.5:
            st.info(t["roi_ok"])
        else:
            st.warning(t["roi_low"])
        st.caption(t["roi_cap"].format(area_name))
    else:
        st.info(t["roi_nodata"])

    # ---- Market charts: one message per chart ----
    st.markdown(f"### {t['ins_head']}")
    ch1, ch2, ch3 = st.tabs(t["ch_tabs"])
    with ch1:
        m = summary.melt("Room Type", ["Median (RM)", "Average (RM)"],
                         var_name="Statistic", value_name="RM")
        m["Statistic"] = m["Statistic"].map(
            {"Median (RM)": t["median_lbl"], "Average (RM)": t["avg_lbl"]})
        fig = px.bar(m, x="Room Type", y="RM", color="Statistic",
                     barmode="group", text_auto=".0f",
                     color_discrete_sequence=["#C8102E", "#F4A6B0"])
        fig.update_traces(textposition="outside")
        fig.update_layout(yaxis_title=t["rm_month"], xaxis_title="",
                          legend_title="", font_size=14, margin=dict(t=30))
        st.plotly_chart(fig, use_container_width=True)
        st.caption(t["ch1_cap"])
    with ch2:
        fig = px.box(df, x="Room Type", y="Price / month (RM)",
                     points="all", color_discrete_sequence=["#C8102E"])
        fig.update_layout(yaxis_title=t["rm_month"], xaxis_title="",
                          font_size=14, margin=dict(t=30))
        st.plotly_chart(fig, use_container_width=True)
        st.caption(t["ch2_cap"])
    with ch3:
        med = df["Price / month (RM)"].median()
        fig = px.histogram(df, x="Price / month (RM)", nbins=25,
                           color_discrete_sequence=["#C8102E"])
        fig.add_vline(x=med, line_dash="dash", line_color="#333",
                      annotation_text=t["median_line"].format(med),
                      annotation_position="top")
        fig.update_layout(yaxis_title=t["n_listings"],
                          xaxis_title=t["rm_month"], font_size=14,
                          bargap=0.05, margin=dict(t=30))
        st.plotly_chart(fig, use_container_width=True)
        st.caption(t["ch3_cap"])

    if not summary.empty:
        best = summary.loc[summary["Units Found"].idxmax()]
        cheap = summary.loc[summary["Average (RM)"].idxmin()]
        med = df["Price / month (RM)"].median()
        sized = df.dropna(subset=["Size (sqft)"])
        psf = ((sized["Price / month (RM)"] / sized["Size (sqft)"]).median()
               if not sized.empty else None)
        msg = t["auto_ins"].format(
            a=area_name, rt=best["Room Type"], n=best["Units Found"],
            f=best["Fair Price (RM)"], c=cheap["Room Type"],
            avg=cheap["Average (RM)"], m=med)
        msg += t["auto_psf"].format(p=psf) if psf else "."
        st.info(msg)

    with st.expander(t["diag"]):
        st.write({"url": url, "pages": res["pages"],
                  "errors": res["errors"], "source": source})


# =================================================== comparison results
def show_comparison():
    cmp = st.session_state.get("cmp")
    if not cmp:
        return
    st.subheader(t["cmp_head"])
    frames = []
    for a, bundle in cmp.items():
        d = pd.DataFrame(bundle["rows"])
        d["Area"] = a
        frames.append(d)
    cdf = pd.concat(frames, ignore_index=True)

    cols = st.columns(len(cmp))
    for col, (a, bundle) in zip(cols, cmp.items()):
        d = pd.DataFrame(bundle["rows"])
        col.metric(a, f"RM {d['Price / month (RM)'].median():,.0f}",
                   f"{len(d)} · {bundle['source']}", delta_color="off")

    comp_rows = []
    for a in cmp:
        d = cdf[cdf["Area"] == a]
        s = summarize(d).set_index("Room Type")["Fair Price (RM)"]
        comp_rows.append({"Area": a, "Units": len(d),
                          "Median (RM)": d["Price / month (RM)"].median(),
                          **{f"{rt} fair (RM)": s.get(rt)
                             for rt in ("Studio", "1BR", "2BR", "3BR")}})
    st.dataframe(pd.DataFrame(comp_rows), use_container_width=True,
                 hide_index=True)

    agg = (cdf.groupby(["Area", "Room Type"])["Price / month (RM)"]
           .median().reset_index())
    fig = px.bar(agg, x="Room Type", y="Price / month (RM)", color="Area",
                 barmode="group", text_auto=".0f", title=t["cmp_chart"])
    fig.update_traces(textposition="outside")
    fig.update_layout(yaxis_title=t["rm_month"], xaxis_title="",
                      font_size=14)
    st.plotly_chart(fig, use_container_width=True)

    pairs = []
    for rt in agg["Room Type"].unique():
        sub = agg[agg["Room Type"] == rt]
        if len(sub) >= 2:
            lo = sub.loc[sub["Price / month (RM)"].idxmin()]
            hi = sub.loc[sub["Price / month (RM)"].idxmax()]
            if hi["Price / month (RM)"] > lo["Price / month (RM)"]:
                save = hi["Price / month (RM)"] - lo["Price / month (RM)"]
                pairs.append(t["cmp_cheaper"].format(
                    t=rt, lo=lo["Area"], d=save, hi=hi["Area"]))
    if pairs:
        st.info("💡 " + " · ".join(pairs))

    stamp = datetime.now().strftime("%Y%m%d")
    st.download_button(t["cmp_dl"], cdf.to_csv(index=False),
                       f"SPEEDHOME_comparison_{stamp}.csv", "text/csv",
                       use_container_width=True)


show_results()
show_comparison()

st.divider()
st.caption(t["footer"])
